import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time 
browser = webdriver.Firefox()
#try:
#wb = openpyxl.load_workbook("final_one.xlsx", read_only=False)
#ws = wb.active
#except:
#wb = openpyxl.Workbook()
#ws = wb.active
#wb = openpyxl.load_workbook("final_one.xlsx", read_only=False)
#ws = wb.active
url_starter = 'http://www.manta.com/'

def each_result(page_url):
    browser.get('http://www.manta.com/')
    time.sleep(20)
    search_input = browser.find_element_by_class_name("form-control navbar-search dropdown-toggle input-sm ng-pristine ng-invalid ng-invalid-required")
    search_input.send_keys('staffing agency usa')
    search_location = browser.find_element_by_id('homeDropdownLocation')
    search_location.click
    search_location.send_keys(Keys.ARROW_DOWN)
    search_location.send_keys(Keys.ARROW_DOWN)
    search_location.send_keys(Keys.RETURN)
    browser.send_keys(Keys.RETURN)
    #browser.get(page_url)
    time.sleep(20)
    media_body = browser.find_elements_by_class_name("media-body")
    for media in media_body:
        media_follow_link = media.find_element_by_tag_name('a')
        link = url_starter + media_follow_link.get_attribute('href')
        company_name = media.find_element_by_tag_name('strong').text
        all_spans = media.find_elements_by_tag_name('span')
        try:
            street_addr = all_spans[0].text
        except:
            street_addr = ""
        try:
            addr_locality = all_spans[1].text
        except:
            addr_locality = ""
        try:
            addr_region = all_spans[2].text
        except:
            addr_region = ""
        try:
            postal_code = all_spans[3].text
        except:
            postal_code = ""
        try:
            phone = media.find_element_by_class_name('hidden-device-xs').text
        except:
            phone = ""
        try:
            website_body = media.find_element_by_class_name('text-right')
            website_anchor = website_body.find_element_by_tag_name('a').text
            if website_anchor == 'Website':
                website = website_anchor.get_attribute('href')
            else:
                website = ""
        except:
            website = ""
        print company_name, street_addr, addr_locality, addr_region, postal_code, phone, website
def each_state(state, state_url):
    browser.get(state_url)
    time.sleep(20)
    #state_name_space = browser.find_element_by_class_name("other_search_criteria")
    #state_name = state_name_space.find_elements_by_tag_name('label')
    #state = state_name[1].text
    #state = "Texas"
    all_posts = browser.find_elements_by_class_name("post")
    for post in all_posts:
        '''try:
            entry_details = post.fnd_element_by_class_name("entry-details")
            email_p = entry_details.find_elements_by_tag_name("p")
            if len(email_p) > 2:
                email = email_p[-1].text
                print email
        except:'''
        email = ""
        entry_title = post.find_element_by_class_name("entry-title")
        company_name = entry_title.find_element_by_tag_name("a").get_attribute('title')
        follow_link = entry_title.find_element_by_tag_name("a").get_attribute('href')
        try:
            phone = post.find_element_by_class_name('phone').text
        except:
            phone = ""
        try:
            addr = post.find_element_by_class_name('address').text
        except:
            addr = ""
        ws.append([company_name, state, phone, addr, email, follow_link])
    wb.save("final_one.xlsx")
def main():
    for i in range(1, 100):
        print i
        source_link = ws.cell(row = i, column = 6).value 
        browser.get(source_link)
        time.sleep(15)
                #browser.send_keys(KE)
        try:
            website = browser.find_element_by_id("website").text
            ws.cell(row = i, column = 7).value = website
            print website
        except:
            print "no website"
              
        try:
            email = browser.find_element_by_class_name("entry-email").text
            ws.cell(row = i, column = 8).value = email
            print email
        except:
            print "no email"
        wb.save("final_one.xlsx") 
    wb.save("final_one.xlsx")



if __name__ == '__main__': 
    each_result('http://www.manta.com/search?search_source=nav&search=staffing+agency+usa&search_location=city+state+or+zip&pt=')   
    #main()
    #for key in data:
    #    each_state(key, data[key])
    #each_state("http://vaporsearchusa.com/?s=+&tag_s=&category=69&articleauthor=&search_custom%5Bpost_city_id%5D=multicity&adv_country=226&adv_zone=3727&adv_city=&search_template=1&post_type=listing&submit=Search")
